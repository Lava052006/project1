import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
FILE_PATH="pg_management_monthly.xlsx"

def create_monthly_sheet(month_name):
    wb=load_workbook(FILE_PATH)
    if month_name in wb.sheetnames:
         print(f"⚠️ Sheet '{month_name}' already exists.")
         return
    
    ws=wb["Customer List"]
    
    master_df=pd.read_excel(FILE_PATH,sheet_name="Customer List")
    new_df=pd.DataFrame(columns=[
            "Tenant Name", "Room No", "Rent", "Elec. Start", "Elec. End",
            "Units Used", "Rate/Unit", "Elec. Bill", "Total Due",
            "Paid On", "Amount Paid", "Balance"
        ])

    for row in ws.iter_rows(min_row=2,values_only=True):
        tenant_name,room_no,rent,checkin,checkout=row
        if checkout is None or checkout=="":
            print(f"Adding {tenant_name} with room {room_no} and rent ₹{rent}")
            new_row = pd.DataFrame([{
                "Tenant Name": tenant_name,
                "Room No": room_no,
                "Rent": rent,
                "Rate/Unit": 10  # Default rate per unit; you can change this
            }])
            new_df=pd.concat([new_df,new_row],ignore_index=True)
             # Add the new sheet to the workbook
            ws_new = wb.create_sheet(title=month_name)

    # Write DataFrame to new sheet
    for r in dataframe_to_rows(new_df, index=False, header=True):
            ws_new.append(r)

    # Save the workbook
    wb.save(FILE_PATH)
    print(f"✅ Monthly sheet '{month_name}' created successfully.")
    



def main():
    choice = input("Enter current month")
    create_monthly_sheet(choice)
if __name__ == "__main__":
    main()   
    
